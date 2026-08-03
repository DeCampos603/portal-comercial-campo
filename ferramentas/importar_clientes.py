#!/usr/bin/env python3
"""
Converte as planilhas de clientes (ativos, em recuperação, inativos) em clientes.json.

Aplica as regras de limpeza de conhecimento/04-dados-de-clientes.md:
  - remove o sufixo redundante "[codigo]" do nome
  - normaliza telefone, CEP e grafia de cidade
  - descarta a coluna "Representante" (valor único em toda a base)
  - deriva grupoEconomico e whatsapp

As três carteiras são opcionais e podem ser importadas juntas ou separadas.
Cliente que aparece em mais de uma fica com a classificação MAIS ativa —
ver ORIGENS abaixo.

Uso:
    python importar_clientes.py --ativos c.xlsx \
        --inativos a.xlsx --recuperacao b.xlsx \
        --saida dados/privado/clientes.json
"""

import argparse
import json
import re
import sys
import unicodedata
import warnings
from collections import Counter, defaultdict
from datetime import datetime, timezone, timedelta
from pathlib import Path

import openpyxl

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# O console do Windows abre em cp1252 e quebra com emoji/acento no relatório.
for _fluxo in (sys.stdout, sys.stderr):
    try:
        _fluxo.reconfigure(encoding="utf-8", errors="replace")
    except (AttributeError, ValueError):
        pass

FUSO_BR = timezone(timedelta(hours=-3))

# Carteiras, em ordem de PRECEDÊNCIA: da mais ativa para a menos.
#
# A ordem não é enfeite. Um cliente pode constar em duas listas — foi
# recuperado e voltou a comprar, mas ainda figura no relatório de inativos.
# Classificá-lo como inativo esconderia justamente quem está comprando.
# Na dúvida, o portal assume a situação mais favorável e deixa o representante
# corrigir; o contrário faria o cliente sumir da lista de quem visitar.
ORIGENS = (
    ("ativos", "ativo"),
    ("recuperacao", "recuperacao"),
    ("inativos", "inativo"),
)

# Colunas (0-indexado, na ordem em que aparecem na Sheet0)
COL_CODIGO, COL_NOME, COL_REPRESENTANTE, COL_CONTATO = 0, 1, 2, 3
COL_FONE, COL_EMAIL, COL_LOGRADOURO, COL_BAIRRO = 4, 5, 6, 7
COL_CIDADE, COL_UF, COL_CEP, COL_STATUS = 8, 9, 10, 11

# Grafia oficial das cidades — a base mistura "SÃO GONÇALO" e "SAO GONCALO".
CIDADES = {
    "RIO DE JANEIRO": "Rio de Janeiro", "SAO GONCALO": "São Gonçalo",
    "NOVA IGUACU": "Nova Iguaçu", "DUQUE DE CAXIAS": "Duque de Caxias",
    "NITEROI": "Niterói", "MACAE": "Macaé", "VOLTA REDONDA": "Volta Redonda",
    "SAO JOAO DE MERITI": "São João de Meriti", "RIO DAS OSTRAS": "Rio das Ostras",
    "ITABORAI": "Itaboraí", "ANGRA DOS REIS": "Angra dos Reis",
    "TERESOPOLIS": "Teresópolis", "MESQUITA": "Mesquita", "NILOPOLIS": "Nilópolis",
    "CAMPOS DOS GOYTACAZES": "Campos dos Goytacazes", "MARICA": "Maricá",
    "CABO FRIO": "Cabo Frio", "ITAGUAI": "Itaguaí", "ARARUAMA": "Araruama",
    "SAQUAREMA": "Saquarema", "RIO BONITO": "Rio Bonito",
    "SAO PEDRO DA ALDEIA": "São Pedro da Aldeia", "BARRA MANSA": "Barra Mansa",
    "BOM JARDIM": "Bom Jardim", "ARMACAO DOS BUZIOS": "Armação dos Búzios",
    "SERRA": "Serra", "ERECHIM": "Erechim",
}

# Prefixos de logradouro que vêm abreviados e inconsistentes
PREFIXOS = [
    (r"^AVENIDA\.?\s+", "Avenida "), (r"^AV\.?\s+", "Avenida "),
    (r"^RUA\.?\s+", "Rua "), (r"^R\.?\s+", "Rua "),
    (r"^ESTRADA\.?\s+", "Estrada "), (r"^EST\.?\s+", "Estrada "),
    (r"^RODOVIA\.?\s+", "Rodovia "), (r"^ROD\.?\s+", "Rodovia "),
    (r"^PRACA\.?\s+", "Praça "), (r"^TRAVESSA\.?\s+", "Travessa "),
]


def sem_acento(texto: str) -> str:
    decomposto = unicodedata.normalize("NFKD", str(texto).upper())
    limpo = "".join(c for c in decomposto if not unicodedata.combining(c))
    return " ".join(limpo.split())


def limpar(valor):
    if valor is None:
        return None
    texto = str(valor).strip()
    return texto or None


def limpar_nome(bruto):
    """'ASPEC ... LTDA [21152]' -> 'ASPEC ... LTDA'. Preserva CNPJ no início (MEIs)."""
    if not bruto:
        return None
    nome = re.sub(r"\s*\[\d+\]\s*$", "", str(bruto).strip())
    nome = re.sub(r"^['\"]+", "", nome)          # apóstrofo residual do Excel
    return " ".join(nome.split()) or None


def normalizar_telefone(bruto):
    """Devolve (formatado, e164). Formato desconhecido é preservado, nunca descartado."""
    if not bruto or not str(bruto).strip():
        return None, None
    digitos = re.sub(r"\D", "", str(bruto))
    if digitos.startswith("0") and len(digitos) > 10:    # DDD escrito como (021)
        digitos = digitos[1:]
    if len(digitos) == 11:
        return f"({digitos[:2]}) {digitos[2:7]}-{digitos[7:]}", f"+55{digitos}"
    if len(digitos) == 10:
        return f"({digitos[:2]}) {digitos[2:6]}-{digitos[6:]}", f"+55{digitos}"
    return str(bruto).strip(), None


def normalizar_cep(bruto):
    digitos = re.sub(r"\D", "", str(bruto or ""))
    return f"{digitos[:5]}-{digitos[5:]}" if len(digitos) == 8 else None


def normalizar_logradouro(bruto):
    if not bruto:
        return None
    texto = " ".join(str(bruto).strip().split())
    texto = re.sub(r"^\.\s*", "", texto)              # alguns começam com "."
    for padrao, troca in PREFIXOS:
        novo = re.sub(padrao, troca, texto, flags=re.IGNORECASE)
        if novo != texto:
            return novo.strip()
    return texto


def titulo(texto):
    """Title Case preservando acento. 'BARRA DA TIJUCA' -> 'Barra da Tijuca'."""
    if not texto:
        return None
    minusculas = {"da", "de", "do", "das", "dos", "e"}
    palavras = str(texto).strip().lower().split()
    return " ".join(
        p if i and p in minusculas else p.capitalize()
        for i, p in enumerate(palavras)
    )


def ler_planilha(caminho: Path, origem: str):
    planilha = openpyxl.load_workbook(caminho, data_only=True)
    aba = planilha[planilha.sheetnames[0]]
    clientes, avisos = [], []

    for numero_linha, linha in enumerate(aba.iter_rows(min_row=2, values_only=True), start=2):
        codigo = limpar(linha[COL_CODIGO])
        if not codigo:
            continue
        codigo = str(int(float(codigo))) if str(codigo).replace(".", "").isdigit() else codigo

        fone, e164 = normalizar_telefone(linha[COL_FONE])
        cep = normalizar_cep(linha[COL_CEP])
        if not cep:
            avisos.append(f"{caminho.name} L{numero_linha}: CEP inválido "
                          f"({linha[COL_CEP]!r}) — cliente ficará sem coordenada.")

        cidade_bruta = limpar(linha[COL_CIDADE]) or ""
        cidade = CIDADES.get(sem_acento(cidade_bruta), titulo(cidade_bruta))

        email = limpar(linha[COL_EMAIL])
        if email:
            email = email.lower()

        clientes.append({
            "codigo": codigo,
            "nome": limpar_nome(linha[COL_NOME]),
            "origem": origem,
            "status": limpar(linha[COL_STATUS]),
            "contato": limpar(linha[COL_CONTATO]),   # 100% vazio na origem
            "telefone": fone,
            "whatsapp": e164,
            "email": email,
            "endereco": {
                "logradouro": normalizar_logradouro(linha[COL_LOGRADOURO]),
                "bairro": titulo(limpar(linha[COL_BAIRRO])),
                "cidade": cidade,
                "uf": limpar(linha[COL_UF]),
                "cep": cep,
            },
            "geo": None,          # preenchido por geocodificar.py
            "notas": None,
            "ultimaVisita": None,
        })

    return clientes, avisos


def derivar_grupo_economico(clientes):
    """Agrupa matriz/filial pela razão social normalizada (FRIOTEC x3, SUL FLUMINENSE x4)."""
    por_nome = defaultdict(list)
    for cliente in clientes:
        if cliente["nome"]:
            por_nome[sem_acento(cliente["nome"])].append(cliente)

    grupos = 0
    for chave, membros in por_nome.items():
        if len(membros) > 1:
            grupos += 1
            for cliente in membros:
                cliente["grupoEconomico"] = chave
    for cliente in clientes:
        cliente.setdefault("grupoEconomico", None)
    return grupos


def main():
    parser = argparse.ArgumentParser(description="Importa as carteiras de clientes.")
    for opcao, _ in ORIGENS:
        parser.add_argument(f"--{opcao}", type=Path)
    parser.add_argument("--saida", type=Path, default=Path("dados/privado/clientes.json"))
    args = parser.parse_args()

    informadas = [(getattr(args, opcao.replace("-", "_")), origem)
                  for opcao, origem in ORIGENS
                  if getattr(args, opcao.replace("-", "_"))]
    if not informadas:
        raise SystemExit("Informe ao menos uma carteira: "
                         + ", ".join(f"--{o}" for o, _ in ORIGENS))

    clientes, avisos, vistos = [], [], {}
    for caminho, origem in informadas:      # já vem na ordem de precedência
        if not caminho.exists():
            raise SystemExit(f"Arquivo não encontrado: {caminho}")

        lidos, avisos_desta = ler_planilha(caminho, origem)
        avisos += avisos_desta

        repetidos = []
        for cliente in lidos:
            anterior = vistos.get(cliente["codigo"])
            if anterior:
                repetidos.append(f"{cliente['codigo']} (fica como {anterior})")
                continue
            vistos[cliente["codigo"]] = origem
            clientes.append(cliente)

        if repetidos:
            avisos.append(
                f"⚠️  {len(repetidos)} cliente(s) de '{origem}' já vieram numa "
                f"carteira mais ativa: {', '.join(repetidos[:5])}"
                + (" …" if len(repetidos) > 5 else "")
            )

    clientes.sort(key=lambda c: c["nome"] or "")
    grupos = derivar_grupo_economico(clientes)

    agora = datetime.now(FUSO_BR)
    args.saida.parent.mkdir(parents=True, exist_ok=True)
    args.saida.write_text(
        json.dumps({
            "versao": agora.strftime("%Y-%m-%d"),
            "atualizadoEm": agora.isoformat(timespec="seconds"),
            "clientes": clientes,
        }, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    # ---- Relatório -------------------------------------------------------
    sem_email = [c for c in clientes if not c["email"]]
    sem_fone = [c for c in clientes if not c["telefone"]]
    sem_contato = [c for c in clientes if not c["contato"]]
    sem_cep = [c for c in clientes if not c["endereco"]["cep"]]
    status = Counter(c["status"] for c in clientes)
    ufs = Counter(c["endereco"]["uf"] for c in clientes)

    print(f"\n{'=' * 58}")
    print(f"CLIENTES IMPORTADOS — {agora:%d/%m/%Y %H:%M}")
    print("=" * 58)
    print(f"Total              {len(clientes)}")
    for _, origem in ORIGENS:
        quantos = sum(1 for c in clientes if c["origem"] == origem)
        if quantos:
            print(f"  {origem:<16} {quantos}")
    print(f"\nStatus             {dict(status)}")
    print(f"UF                 {dict(ufs)}")
    print(f"Cidades            {len({c['endereco']['cidade'] for c in clientes})}")
    print(f"Grupos econômicos  {grupos} (matriz/filial)")
    print()
    print(f"⚠️  Sem contato      {len(sem_contato)}  (coluna vazia na origem)")
    print(f"⚠️  Sem e-mail       {len(sem_email)}")
    print(f"⚠️  Sem telefone     {len(sem_fone)}")
    if sem_cep:
        print(f"🔴 SEM CEP VÁLIDO   {len(sem_cep)} — não entrarão no mapa")

    if avisos:
        print(f"\nAvisos ({len(avisos)}):")
        for aviso in avisos[:15]:
            print(f"  • {aviso}")
        if len(avisos) > 15:
            print(f"  … e mais {len(avisos) - 15}")

    print(f"\n✅ Gravado em {args.saida}")
    print("   Próximo passo: python ferramentas/geocodificar.py")
    return 0


if __name__ == "__main__":
    sys.exit(main())
