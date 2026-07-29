#!/usr/bin/env python3
"""
Inspeciona a estrutura de uma planilha ANTES de importar.

A tabela de preços é atualizada toda semana. Rode isto sempre antes de importar:
se o layout mudou (aba renomeada, coluna trocada de lugar), a importação geraria
dado errado em silêncio — exatamente o que este projeto existe para eliminar.

Uso:
    python inspecionar_planilha.py "Tabela de precos.xlsx"
    python inspecionar_planilha.py arquivo.xlsx --linhas 30
"""

import argparse
import sys
import warnings
from pathlib import Path

import openpyxl

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

for _fluxo in (sys.stdout, sys.stderr):
    try:
        _fluxo.reconfigure(encoding="utf-8", errors="replace")
    except (AttributeError, ValueError):
        pass

# O que esperamos encontrar, segundo conhecimento/03-motor-de-cotacoes.md
ESPERADO = {
    "abas": [
        "PREÇOS OUTROS ESTADOS - CLIENTE",
        "PREÇOS OUTROS ESTADOS-COMISSÃO",
        "Planilha1",
    ],
    "linha_cabecalho": 19,
    "colunas": ["Item", "code", "Código Sigma", "Compatibilidade / Descrição",
                "Qtd.", "IPI", "ST", "Valor Unitario"],
}


def truncar(valor, tamanho=22):
    if valor is None:
        return ""
    texto = str(valor).replace("\n", " ").strip()
    return texto[:tamanho - 1] + "…" if len(texto) > tamanho else texto


def main():
    parser = argparse.ArgumentParser(description="Inspeciona a estrutura de um .xlsx")
    parser.add_argument("xlsx", type=Path)
    parser.add_argument("--linhas", type=int, default=25,
                        help="Quantas linhas exibir por aba (padrão: 25)")
    args = parser.parse_args()

    if not args.xlsx.exists():
        raise SystemExit(f"Arquivo não encontrado: {args.xlsx}")

    planilha = openpyxl.load_workbook(args.xlsx, data_only=True)

    print("=" * 70)
    print(f"ARQUIVO  {args.xlsx.name}")
    print(f"ABAS     {planilha.sheetnames}")
    print("=" * 70)

    for aba in planilha.worksheets:
        print(f"\n┌─ ABA '{aba.title}'  ({aba.max_row} linhas × {aba.max_column} colunas)")
        for numero, linha in enumerate(aba.iter_rows(values_only=True), 1):
            if numero > args.linhas:
                print(f"│  … truncado em {args.linhas} de {aba.max_row} linhas")
                break
            celulas = [truncar(c) for c in linha]
            if not any(celulas):
                continue
            print(f"│ L{numero:>3} │ " + " │ ".join(celulas))

    # ---- Conferência contra o layout esperado ----------------------------
    print("\n" + "=" * 70)
    print("CONFERÊNCIA DO LAYOUT")
    print("=" * 70)

    problemas = []

    for esperada in ESPERADO["abas"]:
        if esperada in planilha.sheetnames:
            print(f"  ✅ Aba '{esperada}'")
        else:
            print(f"  🔴 Aba '{esperada}' NÃO ENCONTRADA")
            problemas.append(f"aba ausente: {esperada}")

    aba_principal = ESPERADO["abas"][0]
    if aba_principal in planilha.sheetnames:
        aba = planilha[aba_principal]
        cabecalho = [
            str(c.value).strip() if c.value else ""
            for c in aba[ESPERADO["linha_cabecalho"]]
        ]
        print(f"\n  Cabeçalho na linha {ESPERADO['linha_cabecalho']}:")
        for i, esperada in enumerate(ESPERADO["colunas"]):
            achada = cabecalho[i] if i < len(cabecalho) else ""
            if achada == esperada:
                print(f"    ✅ col {chr(65 + i)}: {esperada}")
            else:
                print(f"    🔴 col {chr(65 + i)}: esperava '{esperada}', achou '{achada}'")
                problemas.append(f"coluna {chr(65 + i)} mudou")

        # Onde os itens terminam de verdade
        ultimo, quantidade = None, 0
        for numero, linha in enumerate(
            aba.iter_rows(min_row=ESPERADO["linha_cabecalho"] + 1, values_only=True),
            start=ESPERADO["linha_cabecalho"] + 1,
        ):
            if isinstance(linha[0], (int, float)):
                ultimo, quantidade = numero, quantidade + 1
        print(f"\n  Itens: {quantidade}, terminando na linha {ultimo}")
        print("  (a importação lê TODOS — a planilha original soma só até a linha 538)")

    print()
    if problemas:
        print(f"🔴 {len(problemas)} divergência(s) no layout.")
        print("   PARE e confira com o usuário antes de importar.")
        return 1

    print("✅ Layout conforme o esperado. Pode importar.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
