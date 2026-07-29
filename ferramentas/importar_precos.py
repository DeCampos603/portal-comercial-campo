#!/usr/bin/env python3
"""
Converte a planilha de preços da Sigma em catalogo.json normalizado.

Corrige os dois bugs conhecidos da planilha original:
  1. SUM(I20:I538) deixa os itens 520 e 521 fora do total -> aqui lemos TODOS os itens.
  2. VLOOKUP(...$A$1:$B$12) faz "Ventilador" retornar #N/A -> aqui lemos as 12 categorias.

Uso:
    python importar_precos.py "Tabela de precos.xlsx" --saida dados/privado/catalogo.json
"""

import argparse
import json
import re
import sys
import warnings
from datetime import datetime, timezone, timedelta
from pathlib import Path

import openpyxl

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# O console do Windows abre em cp1252 e quebra com emoji/acento no relatório.
for _fluxo in (sys.stdout, sys.stderr):
    try:
        _fluxo.reconfigure(encoding="utf-8", errors="replace")
    except (AttributeError, ValueError):
        pass

FUSO_BR = timezone(timedelta(hours=-3))

ABA_CLIENTE = "PREÇOS OUTROS ESTADOS - CLIENTE"
ABA_COMISSAO = "PREÇOS OUTROS ESTADOS-COMISSÃO"
ABA_CATEGORIAS = "Planilha1"

LINHA_CABECALHO = 19          # os itens começam na linha seguinte
COL = {                       # 1-indexado, como o openpyxl
    "item": 1, "codigo_fabricante": 2, "codigo_sigma": 3, "descricao": 4,
    "qtd": 5, "ipi": 6, "st": 7, "valor_unitario": 8,
    # catálogo lateral (L=código limpo, M='CÓDIGO - DESCRIÇÃO', N=grupo, O=saldo)
    "lat_codigo": 12, "lat_descricao": 13, "lat_grupo": 14, "lat_saldo": 15,
    # aba COMISSÃO
    "categoria": 11,
}

# Limiares de estoque, extraídos da formatação condicional da planilha:
#   vermelho: INDEX($O..., MATCH($C..., $L..., 0)) < 6
#   amarelo : > 6 e < 200
# ⚠️ A regra original deixa o saldo EXATAMENTE 6 sem cor nenhuma (nem <6, nem >6).
#    Aqui o 6 entra em 'baixo' — a lacuna é bug, não regra.
SALDO_SEM_ESTOQUE = 6         # abaixo disso: vermelho
SALDO_BAIXO = 200             # abaixo disso: amarelo


def classificar_estoque(saldo):
    """Reproduz o semáforo da planilha. None quando não há dado de saldo."""
    if saldo is None:
        return None
    if saldo < SALDO_SEM_ESTOQUE:
        return "sem_estoque"      # 🔴
    if saldo < SALDO_BAIXO:
        return "baixo"            # 🟡 (inclui o 6, que a planilha deixa de fora)
    return "ok"


def limpar(valor):
    """Texto aparado, ou None se vazio."""
    if valor is None:
        return None
    texto = str(valor).strip()
    return texto or None


def para_centavos(valor):
    """Converte para centavos inteiros. Devolve None (nunca 0) se não for número."""
    if valor is None or isinstance(valor, str) and not valor.strip():
        return None
    try:
        numero = float(valor)
    except (TypeError, ValueError):
        return None                      # cobre '#N/A' e afins
    if numero != numero:                 # NaN
        return None
    return round(numero * 100)


def ler_categorias(planilha):
    """Tabela categoria -> alíquota. Lê TODAS as linhas (o VLOOKUP original para na 12)."""
    if ABA_CATEGORIAS not in planilha.sheetnames:
        raise SystemExit(f"Aba '{ABA_CATEGORIAS}' não encontrada.")

    aba = planilha[ABA_CATEGORIAS]
    comissoes = {}
    for categoria, aliquota in aba.iter_rows(min_row=2, max_col=2, values_only=True):
        nome = limpar(categoria)
        if not nome:
            continue
        try:
            comissoes[nome] = round(float(aliquota), 6)
        except (TypeError, ValueError):
            print(f"  ⚠️  Categoria '{nome}' com alíquota inválida: {aliquota!r}")
    return comissoes


def ler_catalogo_lateral(aba):
    """
    Catálogo lateral: L=código | M='CÓDIGO - DESCRIÇÃO' | N=grupo | O=saldo.

    A chave é a coluna L (código limpo) — é a mesma que a formatação condicional
    da planilha usa no MATCH. Se L estiver vazia, cai para o prefixo de M.
    """
    lateral = {}
    for linha in aba.iter_rows(min_row=LINHA_CABECALHO + 1, values_only=True):
        codigo = limpar(linha[COL["lat_codigo"] - 1])
        if not codigo:
            bruto = limpar(linha[COL["lat_descricao"] - 1])
            if not bruto:
                continue
            codigo = bruto.split(" - ")[0].strip()
        if not codigo:
            continue

        saldo = linha[COL["lat_saldo"] - 1]
        saldo = int(saldo) if isinstance(saldo, (int, float)) else None
        lateral[codigo] = {
            "grupo": limpar(linha[COL["lat_grupo"] - 1]),
            "saldo": saldo,
            "statusEstoque": classificar_estoque(saldo),
        }
    return lateral


def ler_categorias_por_item(planilha):
    """Código Sigma -> categoria, a partir da aba COMISSÃO."""
    if ABA_COMISSAO not in planilha.sheetnames:
        print(f"  ⚠️  Aba '{ABA_COMISSAO}' ausente — itens ficarão sem categoria.")
        return {}

    aba = planilha[ABA_COMISSAO]
    por_item = {}
    for linha in aba.iter_rows(min_row=LINHA_CABECALHO + 1, values_only=True):
        # Só linhas de item de verdade: coluna A numérica descarta o rodapé.
        if not isinstance(linha[COL["item"] - 1], (int, float)):
            continue
        codigo = limpar(linha[COL["codigo_sigma"] - 1])
        if codigo:
            por_item[codigo] = limpar(linha[COL["categoria"] - 1])
    return por_item


def importar(caminho_xlsx: Path):
    planilha = openpyxl.load_workbook(caminho_xlsx, data_only=True)

    if ABA_CLIENTE not in planilha.sheetnames:
        raise SystemExit(
            f"Aba '{ABA_CLIENTE}' não encontrada.\n"
            f"Abas disponíveis: {planilha.sheetnames}\n"
            "O layout da planilha mudou — pare e confira antes de importar."
        )

    aba = planilha[ABA_CLIENTE]
    comissoes = ler_categorias(planilha)
    lateral = ler_catalogo_lateral(aba)
    categorias = ler_categorias_por_item(planilha)

    itens, avisos = [], []
    vistos = set()

    for numero_linha, linha in enumerate(
        aba.iter_rows(min_row=LINHA_CABECALHO + 1, values_only=True),
        start=LINHA_CABECALHO + 1,
    ):
        # A coluna "Item" numérica é o que separa item real do rodapé
        # (OBSERVAÇÕES, Dados para Faturamento, Responsável pela Aprovação).
        if not isinstance(linha[COL["item"] - 1], (int, float)):
            continue

        codigo_sigma = limpar(linha[COL["codigo_sigma"] - 1])
        if not codigo_sigma:
            avisos.append(f"L{numero_linha}: item sem Código Sigma — ignorado.")
            continue

        if codigo_sigma in vistos:
            avisos.append(f"L{numero_linha}: Código Sigma duplicado '{codigo_sigma}'.")
            continue
        vistos.add(codigo_sigma)

        valor = para_centavos(linha[COL["valor_unitario"] - 1])
        if valor is None:
            avisos.append(f"L{numero_linha}: '{codigo_sigma}' sem preço válido.")

        # O IPI vem com ruído de ponto flutuante (0.052000000000000005).
        ipi_bruto = linha[COL["ipi"] - 1]
        try:
            ipi = round(float(ipi_bruto), 4)
        except (TypeError, ValueError):
            ipi = 0.0
            avisos.append(f"L{numero_linha}: '{codigo_sigma}' com IPI inválido ({ipi_bruto!r}).")

        categoria = categorias.get(codigo_sigma)
        if categoria and categoria not in comissoes:
            avisos.append(
                f"L{numero_linha}: '{codigo_sigma}' na categoria '{categoria}', "
                "que não existe na tabela de comissões."
            )

        extra = lateral.get(codigo_sigma, {})

        itens.append({
            "codigoSigma": codigo_sigma,
            "codigoFabricante": limpar(linha[COL["codigo_fabricante"] - 1]),
            "descricao": limpar(linha[COL["descricao"] - 1]),
            "valorUnitarioCentavos": valor,
            "ipi": ipi,
            "st": str(limpar(linha[COL["st"] - 1]) or "").lower() == "sim",
            "categoria": categoria,
            "grupo": extra.get("grupo"),
            "saldo": extra.get("saldo"),
            # 'sem_estoque' (🔴) | 'baixo' (🟡) | 'ok' | None
            "statusEstoque": extra.get("statusEstoque"),
        })

    return itens, comissoes, avisos, lateral


def main():
    parser = argparse.ArgumentParser(description="Importa a tabela de preços da Sigma.")
    parser.add_argument("xlsx", type=Path, help="Caminho da planilha .xlsx")
    parser.add_argument("--saida", type=Path, default=Path("dados/privado/catalogo.json"))
    args = parser.parse_args()

    if not args.xlsx.exists():
        raise SystemExit(f"Arquivo não encontrado: {args.xlsx}")

    print(f"Lendo {args.xlsx.name}…")
    itens, comissoes, avisos, lateral = importar(args.xlsx)

    agora = datetime.now(FUSO_BR)
    catalogo = {
        "versao": agora.strftime("%Y-%m-%d"),
        "atualizadoEm": agora.isoformat(timespec="seconds"),
        "origem": args.xlsx.name,
        "comissoes": comissoes,
        "itens": itens,
    }

    args.saida.parent.mkdir(parents=True, exist_ok=True)
    args.saida.write_text(
        json.dumps(catalogo, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    # ---- Relatório -------------------------------------------------------
    sem_grupo = [i for i in itens if not i["grupo"]]
    sem_preco = [i for i in itens if i["valorUnitarioCentavos"] is None]
    com_st = [i for i in itens if i["st"]]

    sem_estoque = [i for i in itens if i["statusEstoque"] == "sem_estoque"]
    baixo = [i for i in itens if i["statusEstoque"] == "baixo"]
    ok = [i for i in itens if i["statusEstoque"] == "ok"]
    sem_dado = [i for i in itens if i["statusEstoque"] is None]
    zerados = [i for i in sem_estoque if i["saldo"] == 0]

    print(f"\n{'=' * 58}")
    print(f"CATÁLOGO IMPORTADO — {agora:%d/%m/%Y %H:%M}")
    print("=" * 58)
    print(f"Itens              {len(itens)}")
    print(f"Categorias         {len(comissoes)}")
    print(f"Grupos             {len({i['grupo'] for i in itens if i['grupo']})}")
    print(f"Catálogo lateral   {len(lateral)} itens de referência")

    print(f"\n--- ESTOQUE (limiares da planilha: <{SALDO_SEM_ESTOQUE} e <{SALDO_BAIXO}) ---")
    print(f"🔴 Sem estoque      {len(sem_estoque):>3}   (dos quais {len(zerados)} zerados)")
    print(f"🟡 Estoque baixo    {len(baixo):>3}")
    print(f"🟢 Estoque ok       {len(ok):>3}")
    print(f"⬜ Sem dado          {len(sem_dado):>3}   (fora do catálogo lateral)")

    if sem_estoque:
        print(f"\n   Sem estoque: "
              f"{', '.join(i['codigoSigma'] for i in sem_estoque[:8])}"
              f"{' …' if len(sem_estoque) > 8 else ''}")

    print()
    if sem_grupo:
        print(f"⚠️  Sem grupo        {len(sem_grupo)}")
    print(f"⚠️  Marcados com ST  {len(com_st)}  (selo visual apenas, sem cálculo)")
    if sem_preco:
        print(f"🔴 SEM PREÇO        {len(sem_preco)}: "
              f"{', '.join(i['codigoSigma'] for i in sem_preco[:5])}")

    if avisos:
        print(f"\nAvisos ({len(avisos)}):")
        for aviso in avisos[:15]:
            print(f"  • {aviso}")
        if len(avisos) > 15:
            print(f"  … e mais {len(avisos) - 15}")

    print(f"\n✅ Gravado em {args.saida}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
